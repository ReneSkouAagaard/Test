import subprocess
import time
from shutil import copyfile
from pathlib import Path

import openpyxl
import uno
from com.sun.star.beans import PropertyValue
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "pensionsalder_model.xlsx"
CALCULATED = Path("/tmp/pension-workbook-calculated.xlsx")
PORT = 2003


def prop(name, value):
    item = PropertyValue()
    item.Name = name
    item.Value = value
    return item


def connect_to_calc():
    local_ctx = uno.getComponentContext()
    resolver = local_ctx.ServiceManager.createInstanceWithContext(
        "com.sun.star.bridge.UnoUrlResolver", local_ctx
    )
    url = f"uno:socket,host=127.0.0.1,port={PORT};urp;StarOffice.ComponentContext"
    last_error = None
    for _ in range(30):
        try:
            return resolver.resolve(url)
        except Exception as exc:
            last_error = exc
            time.sleep(0.5)
    raise RuntimeError(f"Could not connect to LibreOffice: {last_error}")


def calculate_with_libreoffice(workbook=WORKBOOK, calculated=CALCULATED):
    cmd = [
        "soffice",
        "--headless",
        f"--accept=socket,host=127.0.0.1,port={PORT};urp;StarOffice.ComponentContext",
        "--norestore",
        "--nofirststartwizard",
    ]
    proc = subprocess.Popen(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    try:
        ctx = connect_to_calc()
        desktop = ctx.ServiceManager.createInstanceWithContext(
            "com.sun.star.frame.Desktop", ctx
        )
        doc = desktop.loadComponentFromURL(
            uno.systemPathToFileUrl(str(workbook)),
            "_blank",
            0,
            (prop("Hidden", True), prop("ReadOnly", False)),
        )
        if doc is None:
            raise RuntimeError(f"Could not open {workbook}")
        doc.calculateAll()
        doc.storeAsURL(
            uno.systemPathToFileUrl(str(calculated)),
            (prop("FilterName", "Calc MS Excel 2007 XML"), prop("Overwrite", True)),
        )
        doc.close(True)
        desktop.terminate()
    finally:
        proc.terminate()
        try:
            proc.wait(timeout=10)
        except subprocess.TimeoutExpired:
            proc.kill()


def assert_numeric(ws, cell):
    value = ws[cell].value
    if not isinstance(value, (int, float)):
        raise AssertionError(f"{ws.title}!{cell} is not numeric after calculation: {value!r}")
    return value


def calculated_j2_with_change(cell, value):
    variant = Path(f"/tmp/pension-workbook-{cell.lower()}-{value}.xlsx")
    calculated = Path(f"/tmp/pension-workbook-{cell.lower()}-{value}-calculated.xlsx")
    copyfile(WORKBOOK, variant)
    wb = openpyxl.load_workbook(variant)
    wb["Model"][cell] = value
    wb.save(variant)
    calculate_with_libreoffice(variant, calculated)
    result = openpyxl.load_workbook(calculated, data_only=True)
    return assert_numeric(result["Aarlig projektion"], "J2")


def main():
    calculate_with_libreoffice()
    formulas = openpyxl.load_workbook(WORKBOOK, data_only=False)
    formula_model = formulas["Model"]
    formula_projection = formulas["Aarlig projektion"]
    if formula_model["L48"].value != "=$B$18":
        raise AssertionError("Model!L48 should reference the progressionsgraense input in Model!B18")
    if formula_model["H48"].value != "=$B$19-G48":
        raise AssertionError("Model!H48 should calculate payout start from reference age in Model!B19")
    if formula_model["B52"].value != "=((3700000-3582782)+(3640000-3566770)+428000)":
        raise AssertionError("Model!B52 should show the explicit investment-property starting equity calculation")
    if formula_model["K52"].value != "=E52*12":
        raise AssertionError("Model!K52 should show the 144,000 kr annual net amortization after company tax")
    if formula_model["P52"].value != '=IF(A52="Investerings-ejendomme",K52/MAX(0.000001,1-C52)*C52,0)':
        raise AssertionError("Model!P52 should gross up the company tax needed for the net amortization")
    if formula_model["O52"].value != 0.05:
        raise AssertionError("Model!O52 should show the 5% property trading cost")
    if formula_model["F52"].value != "=$B$17":
        raise AssertionError("Model!F52 should take the property holding period from Model!B17")
    if formula_projection["J2"].value != '=IF(A2="","",AO2)':
        raise AssertionError("Aarlig projektion!J2 should stay a short helper-cell reference")
    if "Model!$B$26*(1-Model!$C$48)" in formula_projection["T2"].value:
        raise AssertionError("Aarlig projektion!T2 should treat property income as already after company tax")
    if "Model!$B$26*(1-Model!$C$48)" in formula_projection["Z2"].value:
        raise AssertionError("Aarlig projektion!Z2 should add property income to holding without taxing it again")
    if "Model!$B$26*(1+Model!$B$13)^" in formula_projection["T2"].value:
        raise AssertionError("Aarlig projektion!T2 should keep property income in real kroner")
    if "Model!$L$48*(1+Model!$B$13)^" in formula_projection["T2"].value:
        raise AssertionError("Aarlig projektion!T2 should keep the dividend threshold in real kroner")
    if "-W2" in formula_projection["T2"].value or "-W2" in formula_projection["U2"].value:
        raise AssertionError("Aarlig projektion!T2/U2 should not subtract property amortization from liquid property payout capacity")
    if "0=Model!$B$17" not in formula_projection["Y2"].value:
        raise AssertionError("Aarlig projektion!Y2 should transfer property value to holding based on Model!B17")
    if "V2*Model!$O$52" not in formula_projection["X2"].value:
        raise AssertionError("Aarlig projektion!X2 should show 5% property trading cost before holding transfer")
    if formulas["Udbetalingsplan"]["G8"].value != '=IF(A8="","",IF(A8>=Model!$B$10,Model!$B$23,0)+IF(0<Model!$B$15,Model!$B$25,0)+I8*(1-Model!$M$48)+J8*(1-Model!$N$48))':
        raise AssertionError("Udbetalingsplan!G8 should use the real-value property helper payout")

    wb = openpyxl.load_workbook(CALCULATED, data_only=True)
    projection = wb["Aarlig projektion"]
    model = wb["Model"]

    helper_start = 19
    helper_width = 24
    last_unmet = f"{get_column_letter(helper_start + 120 * helper_width + 21)}2"
    last_nominal_unmet = f"{get_column_letter(helper_start + 120 * helper_width + 23)}2"
    for cell in ["C2", "D2", "F2", "G2", "H2", "J2", "K2", "P2", "Q2", "R2", "S2", "V2", "W2", "X2", "Y2", "AI2", "AM2", last_unmet, last_nominal_unmet]:
        assert_numeric(projection, cell)
    for cell in ["B18", "B19", "B34", "B35", "B36", "B38", "B39", "B40", "B52", "B56", "B57", "B58", "B59", "B60"]:
        assert_numeric(model, cell)
    optimal_age = assert_numeric(model, "B28")
    if optimal_age < model["B5"].value or optimal_age > model["B4"].value:
        raise AssertionError(f"Optimal age is outside the modeled age range: {optimal_age!r}")
    assert_numeric(model, "B29")
    assert_numeric(model, "B30")

    life_offset = int(model["B4"].value - projection["A2"].value)
    end_cols = [helper_start + life_offset * helper_width + i for i in range(16, 21)]
    expected_h2 = sum(projection[f"{get_column_letter(col)}2"].value for col in end_cols)
    actual_h2 = assert_numeric(projection, "H2")
    if abs(actual_h2 - expected_h2) > 1:
        raise AssertionError(f"Aarlig projektion!H2 should use balances at life expectancy, got {actual_h2}, expected {expected_h2}")

    baseline_j2 = assert_numeric(projection, "J2")
    if calculated_j2_with_change("B16", 10_000) == baseline_j2:
        raise AssertionError("Changing Model!B16 did not change Aarlig projektion!J2")
    if abs(calculated_j2_with_change("B13", 0.05) - baseline_j2) > 1:
        raise AssertionError("Changing nominal inflation should not change real-value Aarlig projektion!J2")

    print("Workbook calculation OK")


if __name__ == "__main__":
    main()
