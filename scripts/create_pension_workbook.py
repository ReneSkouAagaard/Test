from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OUTPUT = Path(__file__).resolve().parents[1] / "pensionsalder_model.xlsx"
MAX_YEARS = 121
HOLDING_DEPOT_NAME = "Midler i holdingselskab"
PROPERTY_DEPOT_NAME = "Investerings-ejendomme"
HOLDING_MODEL_ROW = 48
PROPERTY_MODEL_ROW = 52

DEPOTS = [
    {"name": HOLDING_DEPOT_NAME, "row": 41, "balance": "=4041000+2476000", "gain_tax": 0.22, "payout_tax": 0.42, "monthly": 0, "years": 0, "start_offset": "=$B$19-$B$5", "payout_years": 120, "low_dividend_limit": "=$B$18", "low_dividend_tax": 0.27, "salary_tax": 0.38},
    {"name": "Midler paa pension", "row": 42, "balance": 1_175_000, "gain_tax": 0.153, "payout_tax": 0.38, "monthly": 6_000, "years": 3, "start_offset": 3, "payout_years": 120, "low_dividend_limit": 0, "low_dividend_tax": 0, "salary_tax": 0},
    {"name": "Frie midler paa aktiedepot", "row": 43, "balance": 107_000, "gain_tax": 0.42, "payout_tax": 0.0, "monthly": 9_000, "years": 3, "start_offset": "=$B$19-$B$5", "payout_years": 120, "low_dividend_limit": 0, "low_dividend_tax": 0, "salary_tax": 0},
    {"name": "Frie midler paa aktiesparekonto", "row": 44, "balance": 201_000, "gain_tax": 0.17, "payout_tax": 0.0, "monthly": 0, "years": 0, "start_offset": "=$B$19-$B$5", "payout_years": 120, "low_dividend_limit": 0, "low_dividend_tax": 0, "salary_tax": 0},
    {"name": PROPERTY_DEPOT_NAME, "row": 45, "balance": "=((3700000-3582782)+(3640000-3566770)+428000)", "gain_tax": 0.22, "payout_tax": 0.0, "monthly": "=144000/12", "years": "=$B$17", "start_offset": "=$B$19-$B$5", "payout_years": 0, "low_dividend_limit": 0, "low_dividend_tax": 0, "salary_tax": 0, "trade_cost": 0.05},
]


def money_style(cell):
    cell.number_format = '#,##0 "kr"'


def pct_style(cell):
    cell.number_format = "0.00%"


def style_header(row):
    for cell in row:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")


def model_ref(cell):
    return f"Model!${cell[0]}${cell[1:]}"


def depot_balance_at_retirement(depot_row, years_cell):
    rate = model_ref(f"J{depot_row}")
    balance = model_ref(f"B{depot_row}")
    annual_payment = model_ref(f"K{depot_row}")
    if depot_row == HOLDING_MODEL_ROW:
        annual_payment = f"({annual_payment}+MAX(0,Model!$B$26))"
    payment_years = model_ref(f"F{depot_row}")
    n = f"MAX(0,{years_cell})" if depot_row == HOLDING_MODEL_ROW else f"MIN(MAX(0,{years_cell}),{payment_years})"
    return (
        f"({balance}*(1+{rate})^{years_cell}+"
        f"IF({n}<=0,0,IF({rate}=0,{annual_payment}*{n},"
        f"{annual_payment}*((1+{rate})^{n}-1)/{rate})))"
    )


def property_balance_at_retirement(depot_row, years_cell):
    balance = model_ref(f"B{depot_row}")
    annual_payment = model_ref(f"K{depot_row}")
    return f"({balance}+IF(MAX(0,{years_cell})<=0,0,{annual_payment}*MAX(0,{years_cell})))"


def contribution_formula(depot_row, years_from_now):
    return f"IF({years_from_now}<{model_ref(f'F{depot_row}')},{model_ref(f'K{depot_row}')},0)"


def property_contribution_formula(depot_row, offset):
    return f"IF({offset}<Model!$B$17,{model_ref(f'K{depot_row}')},0)"


def available_formula(depot_row, future_age):
    start_age = model_ref(f"H{depot_row}")
    payout_years = model_ref(f"I{depot_row}")
    return f"AND({future_age}>={start_age},{future_age}<({start_age}+{payout_years}))"


def main():
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    ws = wb.active
    ws.title = "Model"
    proj = wb.create_sheet("Aarlig projektion")
    plan = wb.create_sheet("Udbetalingsplan")

    input_fill = PatternFill("solid", fgColor="FFF2CC")

    ws["A1"] = "Selvpensionsmodel"
    ws["A1"].font = Font(size=18, bold=True)
    ws["A3"] = "Input"
    ws["A3"].font = Font(size=13, bold=True)

    inputs = [
        ("Levealder", 95, "aar"),
        ("Nuvaerende alder", 35, "aar"),
        ("Aarligt forbrug efter skat", 700_000, "kr i dag"),
        ("Lavere aarligt forbrug i sidste aar", 400_000, "kr i dag"),
        ("Antal sidste aar med lavere forbrug", 15, "aar"),
        ("Aarligt afkast REAL foer skat", 0.08, "%"),
        ("Folkepension start", 73, "aar"),
        ("Folkepension foer skat", "=7544*2", "kr pr maaned i dag"),
        ("Skat paa oevrige indtaegter", 0.38, "%"),
        ("Inflation til nominelle visninger", 0.02, "%"),
        ("Bijob efter selvpension foer skat", 30_000, "kr pr maaned i dag"),
        ("Bijob varighed efter selvpension", 5, "aar"),
        ("Investeringsejendomme efter selskabsskat", "=(149000+116000)/12", "kr pr maaned i dag, foer personlig skat"),
        ("Investeringsejendomme varighed efter selvpension", 30, "aar"),
        ("Progressionsgraense aktieindkomst", 158_800, "kr i dag, 2 personer"),
        ("Referencealder for depotudbetalinger", 73, "aar"),
    ]

    for row, (label, value, unit) in enumerate(inputs, start=4):
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        ws[f"C{row}"] = unit
        ws[f"B{row}"].fill = input_fill
        if "%" in unit:
            pct_style(ws[f"B{row}"])
        elif "kr" in unit:
            money_style(ws[f"B{row}"])

    ws["A21"] = "Beregninger"
    ws["A21"].font = Font(size=13, bold=True)
    ws["D11"] = "For 2 personer"

    derived = {
        "A22": "Folkepension foer skat, aarligt",
        "B22": "=B11*12",
        "A23": "Folkepension efter skat, aarligt",
        "B23": "=B22*(1-B12)",
        "A24": "Bijob foer skat, aarligt",
        "B24": "=B14*12",
        "A25": "Bijob efter skat, aarligt",
        "B25": "=B24*(1-B12)",
        "A26": "Investeringsejendomme efter selskabsskat, aarligt",
        "B26": "=B16*12",
        "A27": "Investeringsejendomme efter personlig skat som loen, aarligt",
        "B27": "=B26*(1-B12)",
        "A28": "Selvpensionsalder",
        "B28": '=IFERROR(INDEX(\'Aarlig projektion\'!A2:A122,MATCH(TRUE,\'Aarlig projektion\'!I2:I122,0)),"Ikke opnaaet")',
        "A29": "Portefolje ved selvpensionsalder",
        "B29": '=IF(ISNUMBER(B28),INDEX(\'Aarlig projektion\'!C2:C122,MATCH(B28,\'Aarlig projektion\'!A2:A122,0)),"")',
        "A30": "Samlet mangel ved selvpensionsalder",
        "B30": '=IF(ISNUMBER(B28),INDEX(\'Aarlig projektion\'!K2:K122,MATCH(B28,\'Aarlig projektion\'!A2:A122,0)),"")',
        "A31": "Vaegtet realafkast efter skat",
        "B31": "=SUMPRODUCT(B48:B51,J48:J51)/SUM(B48:B51)",
    }
    for cell, value in derived.items():
        ws[cell] = value
    for row in [22, 23, 24, 25, 26, 27, 29, 30]:
        money_style(ws[f"B{row}"])
    pct_style(ws["B31"])

    ws["A32"] = "Oversigt pr. mulig selvpensionsalder"
    ws["A32"].font = Font(size=13, bold=True)
    ws["A33"] = "Alder"
    ws["A34"] = "Portefolje"
    ws["A35"] = "Behov for selvpension"
    ws["A36"] = "Likviditetsmangel"
    ws["A37"] = "Foerste alder med likviditetsmangel"
    ws["A38"] = "Portefolje nominelt"
    ws["A39"] = "Behov for selvpension nominelt"
    ws["A40"] = "Likviditetsmangel nominelt"
    for col in range(2, MAX_YEARS + 2):
        col_letter = get_column_letter(col)
        ws[f"{col_letter}33"] = f'=IF($B$5+COLUMN()-2<=$B$4,$B$5+COLUMN()-2,"")'
        ws[f"{col_letter}34"] = f'=IF({col_letter}$33="","",INDEX(\'Aarlig projektion\'!$C:$C,MATCH({col_letter}$33,\'Aarlig projektion\'!$A:$A,0)))'
        ws[f"{col_letter}35"] = f'=IF({col_letter}$33="","",INDEX(\'Aarlig projektion\'!$J:$J,MATCH({col_letter}$33,\'Aarlig projektion\'!$A:$A,0)))'
        ws[f"{col_letter}36"] = f'=IF({col_letter}$33="","",INDEX(\'Aarlig projektion\'!$K:$K,MATCH({col_letter}$33,\'Aarlig projektion\'!$A:$A,0)))'
        ws[f"{col_letter}37"] = f'=IF({col_letter}$33="","",INDEX(\'Aarlig projektion\'!$O:$O,MATCH({col_letter}$33,\'Aarlig projektion\'!$A:$A,0)))'
        ws[f"{col_letter}38"] = f'=IF({col_letter}$33="","",INDEX(\'Aarlig projektion\'!$Q:$Q,MATCH({col_letter}$33,\'Aarlig projektion\'!$A:$A,0)))'
        ws[f"{col_letter}39"] = f'=IF({col_letter}$33="","",INDEX(\'Aarlig projektion\'!$R:$R,MATCH({col_letter}$33,\'Aarlig projektion\'!$A:$A,0)))'
        ws[f"{col_letter}40"] = f'=IF({col_letter}$33="","",INDEX(\'Aarlig projektion\'!$P:$P,MATCH({col_letter}$33,\'Aarlig projektion\'!$A:$A,0)))'
        for overview_row in [34, 35, 36, 38, 39, 40]:
            money_style(ws[f"{col_letter}{overview_row}"])
        ws.column_dimensions[col_letter].width = 14
    style_header(ws[33])
    ws["A46"] = "Depotinput"
    ws["A46"].font = Font(size=13, bold=True)
    depot_headers = [
        "Depot",
        "Startsaldo",
        "Skat paa vaerdistigning",
        "Skat paa udbetaling",
        "Maanedlig indbetaling",
        "Indbetaling varighed aar",
        "Udbetaling starter aar foer referencealder",
        "Udbetaling startalder",
        "Udbetaling varighed aar",
        "Aarligt realafkast efter skat",
        "Aarlig indbetaling",
        "Holding udbytte lav skat max",
        "Holding udbytteskat lav",
        "Holding loenskat",
        "Handelsomkostning ved ejendomssalg",
        "Selskabsskat af aarligt afdrag",
        "Netto aarligt afdrag efter selskabsskat",
    ]
    for col, header in enumerate(depot_headers, start=1):
        ws.cell(47, col, header)
    style_header(ws[47])

    for depot in DEPOTS:
        row = depot["row"] + 7
        ws[f"A{row}"] = depot["name"]
        ws[f"B{row}"] = depot["balance"]
        ws[f"C{row}"] = depot["gain_tax"]
        ws[f"D{row}"] = depot["payout_tax"]
        ws[f"E{row}"] = depot["monthly"]
        ws[f"F{row}"] = depot["years"]
        ws[f"G{row}"] = depot["start_offset"]
        ws[f"H{row}"] = f"=$B$19-G{row}"
        ws[f"I{row}"] = depot["payout_years"]
        ws[f"J{row}"] = f"=$B$9*(1-C{row})"
        ws[f"K{row}"] = f"=E{row}*12"
        if depot["name"] == PROPERTY_DEPOT_NAME:
            ws[f"J{row}"] = 0
            ws[f"K{row}"] = f"=E{row}*12"
        ws[f"L{row}"] = depot["low_dividend_limit"]
        ws[f"M{row}"] = depot["low_dividend_tax"]
        ws[f"N{row}"] = depot["salary_tax"]
        ws[f"O{row}"] = depot.get("trade_cost", 0)
        ws[f"P{row}"] = f'=IF(A{row}="{PROPERTY_DEPOT_NAME}",K{row}/MAX(0.000001,1-C{row})*C{row},0)'
        ws[f"Q{row}"] = f'=IF(A{row}="{PROPERTY_DEPOT_NAME}",K{row},0)'
        for cell in [f"B{row}", f"E{row}", f"K{row}", f"L{row}", f"P{row}", f"Q{row}"]:
            money_style(ws[cell])
        for cell in [f"C{row}", f"D{row}", f"J{row}", f"M{row}", f"N{row}", f"O{row}"]:
            pct_style(ws[cell])
        for col in range(2, 18):
            ws.cell(row, col).fill = input_fill

    ws["A55"] = "Ejendomsdepot beregning"
    ws["A55"].font = Font(size=13, bold=True)
    property_row = PROPERTY_MODEL_ROW
    ws["A56"] = "Startvaerdi investerings-ejendomme"
    ws["B56"] = f"=B{property_row}"
    ws["C56"] = "=(3700000-3582782)+(3640000-3566770)+428000"
    ws["A57"] = "Netto aarligt afdrag efter selskabsskat"
    ws["B57"] = f"=K{property_row}"
    ws["A58"] = "Selskabsskat af afdrag"
    ws["B58"] = f"=P{property_row}"
    ws["A59"] = "Bruttoresultat brugt paa afdrag"
    ws["B59"] = f"=B57+B58"
    ws["A60"] = "Handelsomkostning ved overflytning til holding"
    ws["B60"] = f"=O{property_row}"
    for cell in ["B56", "B57", "B58", "B59"]:
        money_style(ws[cell])
    pct_style(ws["B60"])

    ws["A62"] = "Antagelser"
    ws["A62"].font = Font(size=13, bold=True)
    assumptions = [
        "Alle hovedbeloeb er i realkroner, fordi afkastinput er realt.",
        "De sidste aar med lavere forbrug regnes baglaens fra levealderen inklusiv levealder.",
        "Bijob og investeringsejendomme antages at starte ved den valgte selvpensionsalder og loebe i de valgte antal aar.",
        "Depotudbetalinger kan kun bruges fra den beregnede startalder og inden for den valgte udbetalingsvarighed.",
        "Udbetalinger prioriteres pensionsdepot, holding, aktiedepot og ASK; resterende saldo investeres fortsat.",
        "Holding udbetales foerst som lavt beskattet udbytte op til den valgte graense og derefter som loen.",
        "Indtjening fra investeringsejendomme og progressionsgraense holdes i realkroner i simuleringen.",
        "Ejendomsdepotet opbygges med nettoafdrag efter selskabsskat som ekstra illikvid formue og overflyttes til holding efter B17 aar minus handelsomkostning.",
        "Holding loen reducerer selskabets skattepligtige vaerdistigning i modellen, saa selskabsskat kun beregnes af afkast ud over loen.",
        "Behov for selvpension i oversigten er et enkelt nutidsvaerdiestimat baseret paa vaegtet realafkast efter skat.",
        "Likviditetsmangel viser, om depoternes udbetalingsregler giver underskud undervejs selv om samlet portefolje er stor nok.",
        "Nominelle overbliksbeloeb fremskrives med inflationsinputtet til den relevante alder.",
        "Foerste alder med likviditetsmangel viser det foerste simulerede aar, hvor depoterne ikke kan daekke nettoforbruget.",
    ]
    for row, text in enumerate(assumptions, start=63):
        ws[f"A{row}"] = text

    for col, width in {"A": 46, "B": 18, "C": 42, "D": 20, "E": 20, "F": 22, "G": 28, "H": 20, "I": 22, "J": 22, "K": 20, "L": 24, "M": 22, "N": 18, "O": 28, "P": 24, "Q": 28}.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A4"

    headers = [
        "Alder",
        "Aar fra nu",
        "Samlet portefolje ved selvpension",
        "Forbrug efter skat i foerste selvpensionsaar",
        "Indkomst efter skat i foerste selvpensionsaar",
        "Nettoforbrug fra depoter i foerste selvpensionsaar",
        "Udbetalingsskat i foerste selvpensionsaar",
        "Slutsaldo ved levealder",
        "Kan gaa paa selvpension",
        "Behov for selvpension",
        "Samlet likviditetsmangel",
        "Folkepension nominelt foer skat",
        "Bijob nominelt foer skat i foerste selvpensionsaar",
        "Investeringsejendomme nominelt efter selskabsskat i foerste selvpensionsaar",
        "Foerste alder med likviditetsmangel",
        "Samlet likviditetsmangel nominelt",
        "Samlet portefolje ved selvpension nominelt",
        "Behov for selvpension nominelt",
    ]
    for col, header in enumerate(headers, start=1):
        proj.cell(1, col, header)
    style_header(proj[1])

    last_row = MAX_YEARS + 1
    helper_start = 19
    helper_width = 24
    helper_need_cols = []
    helper_unmet_cols = []
    helper_nominal_unmet_cols = []
    helper_end_total_cols = []

    depot_rows_by_name = {depot["name"]: depot["row"] + 7 for depot in DEPOTS}
    holding_row = depot_rows_by_name[HOLDING_DEPOT_NAME]
    pension_row = depot_rows_by_name["Midler paa pension"]
    aktie_row = depot_rows_by_name["Frie midler paa aktiedepot"]
    ask_row = depot_rows_by_name["Frie midler paa aktiesparekonto"]
    property_row = depot_rows_by_name[PROPERTY_DEPOT_NAME]
    end_offsets = {
        holding_row: 16,
        pension_row: 17,
        aktie_row: 18,
        ask_row: 19,
        property_row: 20,
    }

    for offset in range(MAX_YEARS):
        col = helper_start + offset * helper_width
        letters = [get_column_letter(col + i) for i in range(helper_width)]
        for i, header in enumerate(
            [
                f"Behov +{offset}",
                f"Ejendom udbytte +{offset}",
                f"Ejendom loen +{offset}",
                f"Start ejendomsdepot +{offset}",
                f"Ejendomsafdrag efter selskabsskat +{offset}",
                f"Ejendom handelsomkostning +{offset}",
                f"Ejendom overflytning til holding +{offset}",
                f"Start holding +{offset}",
                f"Start pension +{offset}",
                f"Start aktiedepot +{offset}",
                f"Start ASK +{offset}",
                f"Gross pension +{offset}",
                f"Holding udbytte +{offset}",
                f"Holding loen +{offset}",
                f"Gross aktiedepot +{offset}",
                f"Gross ASK +{offset}",
                f"End holding +{offset}",
                f"End pension +{offset}",
                f"End aktiedepot +{offset}",
                f"End ASK +{offset}",
                f"End ejendomsdepot +{offset}",
                f"Mangel +{offset}",
                f"PV behov rest +{offset}",
                f"Mangel nominel +{offset}",
            ],
            start=col,
        ):
            proj.cell(1, i, header)
            proj.column_dimensions[get_column_letter(i)].hidden = True
        helper_need_cols.append(letters[0])
        helper_unmet_cols.append(letters[21])
        helper_nominal_unmet_cols.append(letters[23])
        helper_end_total_cols.append(letters[16:21])

    first_helper_col = get_column_letter(helper_start)
    last_helper_col = get_column_letter(helper_start + MAX_YEARS * helper_width - 1)

    for row in range(2, last_row + 1):
        proj[f"A{row}"] = f'=IF(Model!$B$5+ROW()-2<=Model!$B$4,Model!$B$5+ROW()-2,"")'
        proj[f"B{row}"] = f'=IF(A{row}="","",A{row}-Model!$B$5)'
        starting_balances = []
        for depot in DEPOTS:
            depot_row = depot["row"] + 7
            if depot["name"] == PROPERTY_DEPOT_NAME:
                starting_balances.append(property_balance_at_retirement(depot_row, f"B{row}"))
            else:
                starting_balances.append(depot_balance_at_retirement(depot_row, f"B{row}"))
        proj[f"C{row}"] = f'=IF(A{row}="","",SUM({",".join(starting_balances)}))'
        proj[f"D{row}"] = f'=IF(A{row}="","",IF(A{row}>Model!$B$4-Model!$B$8,Model!$B$7,Model!$B$6))'
        first_property_dividend = get_column_letter(helper_start + 1)
        first_property_salary = get_column_letter(helper_start + 2)
        proj[f"E{row}"] = (
            f'=IF(A{row}="","",IF(A{row}>=Model!$B$10,Model!$B$23,0)+'
            f'IF(Model!$B$15>0,Model!$B$25,0)+'
            f'{first_property_dividend}{row}*(1-Model!$M$48)+{first_property_salary}{row}*(1-Model!$N$48))'
        )
        proj[f"F{row}"] = f'=IF(A{row}="","",MAX(0,D{row}-E{row}))'

        first_gross_col = get_column_letter(helper_start + 11)
        last_gross_col = get_column_letter(helper_start + 15)
        unmet_cells = ",".join(f"{col}{row}" for col in helper_unmet_cols)
        life_offset = f"MIN({MAX_YEARS - 1},MAX(0,Model!$B$4-A{row}))"
        helper_range = f"{first_helper_col}{row}:{last_helper_col}{row}"
        proj[f"G{row}"] = f'=IF(A{row}="","",MAX(0,SUM({first_gross_col}{row}:{last_gross_col}{row})-F{row}))'
        proj[f"H{row}"] = (
            f'=IF(A{row}="","",SUM('
            f'INDEX({helper_range},1,{life_offset}*{helper_width}+17),'
            f'INDEX({helper_range},1,{life_offset}*{helper_width}+18),'
            f'INDEX({helper_range},1,{life_offset}*{helper_width}+19),'
            f'INDEX({helper_range},1,{life_offset}*{helper_width}+20),'
            f'INDEX({helper_range},1,{life_offset}*{helper_width}+21)))'
        )
        proj[f"I{row}"] = f'=IF(A{row}="","",SUM({unmet_cells})=0)'
        proj[f"J{row}"] = f'=IF(A{row}="","",{get_column_letter(helper_start + 22)}{row})'
        proj[f"K{row}"] = f'=IF(A{row}="","",SUM({unmet_cells}))'
        proj[f"L{row}"] = f'=IF(A{row}="","",IF(A{row}>=Model!$B$10,Model!$B$22*(1+Model!$B$13)^B{row},0))'
        proj[f"M{row}"] = f'=IF(A{row}="","",IF(Model!$B$15>0,Model!$B$24*(1+Model!$B$13)^B{row},0))'
        proj[f"N{row}"] = f'=IF(A{row}="","",IF(Model!$B$17>0,Model!$B$26*(1+Model!$B$13)^B{row},0))'
        first_shortage_age = ",".join(
            f"IF({col}{row}>0,A{row}+{offset},999)"
            for offset, col in enumerate(helper_unmet_cols)
        )
        nominal_unmet_cells = ",".join(f"{col}{row}" for col in helper_nominal_unmet_cols)
        proj[f"O{row}"] = f'=IF(A{row}="","",IF(K{row}=0,"",MIN({first_shortage_age})))'
        proj[f"P{row}"] = f'=IF(A{row}="","",SUM({nominal_unmet_cells}))'
        proj[f"Q{row}"] = f'=IF(A{row}="","",C{row}*(1+Model!$B$13)^B{row})'
        proj[f"R{row}"] = f'=IF(A{row}="","",J{row}*(1+Model!$B$13)^B{row})'

        for offset in range(MAX_YEARS):
            col = helper_start + offset * helper_width
            need_col, property_dividend, property_salary, property_start, property_amortization, property_trade_cost, property_transfer, holding_start, pension_start, aktie_start, ask_start = [get_column_letter(col + i) for i in range(11)]
            gross_pension, holding_dividend, holding_salary, gross_aktie, gross_ask = [get_column_letter(col + i) for i in range(11, 16)]
            end_holding, end_pension, end_aktie, end_ask, end_property, unmet_col, pv_col, nominal_unmet_col = [get_column_letter(col + i) for i in range(16, 24)]
            future_age = f"(A{row}+{offset})"
            years_from_now = f"(B{row}+{offset})"
            target = f"IF({future_age}>Model!$B$4-Model!$B$8,Model!$B$7,Model!$B$6)"
            public_income = f"IF({future_age}>=Model!$B$10,Model!$B$23,0)"
            job_income = f"IF({offset}<Model!$B$15,Model!$B$25,0)"
            proj[f"{need_col}{row}"] = f'=IF(A{row}="","",IF({future_age}>Model!$B$4,0,MAX(0,{target}-{public_income}-{job_income})))'

            start_cells = {
                depot_rows_by_name["Midler i holdingselskab"]: holding_start,
                depot_rows_by_name["Midler paa pension"]: pension_start,
                depot_rows_by_name["Frie midler paa aktiedepot"]: aktie_start,
                depot_rows_by_name["Frie midler paa aktiesparekonto"]: ask_start,
                property_row: property_start,
            }
            end_cells = {
                depot_rows_by_name["Midler i holdingselskab"]: end_holding,
                depot_rows_by_name["Midler paa pension"]: end_pension,
                depot_rows_by_name["Frie midler paa aktiedepot"]: end_aktie,
                depot_rows_by_name["Frie midler paa aktiesparekonto"]: end_ask,
                property_row: end_property,
            }
            gross_cells = {
                ask_row: gross_ask,
                aktie_row: gross_aktie,
                pension_row: gross_pension,
            }

            for depot in DEPOTS:
                depot_row = depot["row"] + 7
                start_cell = start_cells[depot_row]
                if offset == 0:
                    if depot["name"] == PROPERTY_DEPOT_NAME:
                        base = property_balance_at_retirement(depot_row, f"B{row}")
                    else:
                        base = depot_balance_at_retirement(depot_row, f"B{row}")
                else:
                    prev_end = get_column_letter(col - helper_width + end_offsets[depot_row])
                    base = f"{prev_end}{row}"
                if depot["name"] == PROPERTY_DEPOT_NAME:
                    contribution = property_contribution_formula(depot_row, offset)
                    proj[f"{property_amortization}{row}"] = f'=IF(A{row}="","",{contribution})'
                    proj[f"{start_cell}{row}"] = f'=IF(A{row}="","",{base}+{property_amortization}{row})'
                else:
                    proj[f"{start_cell}{row}"] = f'=IF(A{row}="","",{base}+{contribution_formula(depot_row, years_from_now)})'

            remaining = f"{need_col}{row}"
            property_gross = f"IF({offset}<Model!$B$17,Model!$B$26,0)"
            property_after_corp = f"MAX(0,{property_gross})"
            low_dividend_limit = model_ref(f"L{holding_row}")
            proj[f"{property_dividend}{row}"] = (
                f'=IF(A{row}="","",MIN({property_after_corp},{low_dividend_limit},'
                f'MAX(0,{remaining})/MAX(0.000001,1-Model!$M$48)))'
            )
            remaining_after_property_dividend = f"MAX(0,{remaining}-{property_dividend}{row}*(1-Model!$M$48))"
            property_salary_capacity = f"MAX(0,{property_gross}/MAX(0.000001,1-Model!$C$48)-{property_dividend}{row}/MAX(0.000001,1-Model!$C$48))"
            proj[f"{property_salary}{row}"] = (
                f'=IF(A{row}="","",MIN({property_salary_capacity},'
                f'MAX(0,{remaining_after_property_dividend})/MAX(0.000001,1-Model!$N$48)))'
            )
            remaining = f"MAX(0,{remaining_after_property_dividend}-{property_salary}{row}*(1-Model!$N$48))"

            pension_available = available_formula(pension_row, future_age)
            pension_tax = model_ref(f"D{pension_row}")
            proj[f"{gross_pension}{row}"] = (
                f'=IF(A{row}="","",IF({pension_available},'
                f'MIN({pension_start}{row},MAX(0,{remaining})/MAX(0.000001,1-{pension_tax})),0))'
            )
            remaining = f"MAX(0,{remaining}-{gross_pension}{row}*(1-{pension_tax}))"

            holding_available = available_formula(holding_row, future_age)
            holding_div_tax = model_ref(f"M{holding_row}")
            holding_salary_tax = model_ref(f"N{holding_row}")
            holding_limit = f"MAX(0,{low_dividend_limit}-{property_dividend}{row})"
            proj[f"{holding_dividend}{row}"] = (
                f'=IF(A{row}="","",IF({holding_available},'
                f'MIN({holding_start}{row},{holding_limit},MAX(0,{remaining})/MAX(0.000001,1-{holding_div_tax})),0))'
            )
            remaining_after_dividend = f"MAX(0,{remaining}-{holding_dividend}{row}*(1-{holding_div_tax}))"
            proj[f"{holding_salary}{row}"] = (
                f'=IF(A{row}="","",IF({holding_available},'
                f'MIN(MAX(0,{holding_start}{row}-{holding_dividend}{row}),'
                f'MAX(0,{remaining_after_dividend})/MAX(0.000001,1-{holding_salary_tax})),0))'
            )
            remaining = f"MAX(0,{remaining_after_dividend}-{holding_salary}{row}*(1-{holding_salary_tax}))"

            for depot_row in [aktie_row, ask_row]:
                gross_cell = gross_cells[depot_row]
                start_cell = start_cells[depot_row]
                tax = model_ref(f"D{depot_row}")
                available = available_formula(depot_row, future_age)
                proj[f"{gross_cell}{row}"] = (
                    f'=IF(A{row}="","",IF({available},'
                    f'MIN({start_cell}{row},MAX(0,{remaining})/MAX(0.000001,1-{tax})),0))'
                )
                remaining = f"MAX(0,{remaining}-{gross_cell}{row}*(1-{tax}))"

            for depot in DEPOTS:
                depot_row = depot["row"] + 7
                start_cell = start_cells[depot_row]
                end_cell = end_cells[depot_row]
                if depot_row == holding_row:
                    gross_cell = f"({holding_dividend}{row}+{holding_salary}{row})"
                    property_transfer_amount = f"{property_transfer}{row}"
                    gross_return = f"MAX(0,{start_cell}{row}+{property_transfer_amount}-{gross_cell})*Model!$B$9"
                    corp_tax = f"MAX(0,{gross_return}-{holding_salary}{row})*Model!$C${holding_row}"
                    proj[f"{end_cell}{row}"] = f'=IF(A{row}="","",MAX(0,{start_cell}{row}+{property_transfer_amount}-{gross_cell})+{gross_return}-{corp_tax})'
                elif depot_row == property_row:
                    trade_cost = f"IF({offset}=Model!$B$17,{start_cell}{row}*Model!$O${property_row},0)"
                    transfer = f"IF({offset}=Model!$B$17,{start_cell}{row}-{property_trade_cost}{row},0)"
                    proj[f"{property_trade_cost}{row}"] = f'=IF(A{row}="","",{trade_cost})'
                    proj[f"{property_transfer}{row}"] = f'=IF(A{row}="","",{transfer})'
                    proj[f"{end_cell}{row}"] = f'=IF(A{row}="","",MAX(0,{start_cell}{row}-{property_trade_cost}{row}-{property_transfer}{row}))'
                else:
                    gross_cell = gross_cells.get(depot_row)
                    if gross_cell is None:
                        raise RuntimeError("Missing gross cell")
                    rate = model_ref(f"J{depot_row}")
                    proj[f"{end_cell}{row}"] = f'=IF(A{row}="","",MAX(0,{start_cell}{row}-{gross_cell}{row})*(1+{rate}))'
            proj[f"{unmet_col}{row}"] = f'=IF(A{row}="","",{remaining})'
            proj[f"{nominal_unmet_col}{row}"] = f'=IF(A{row}="","",{unmet_col}{row}*(1+Model!$B$13)^{years_from_now})'
            adjusted_need = (
                f"MAX(0,{need_col}{row}-{property_dividend}{row}*(1-Model!$M$48)-"
                f"{property_salary}{row}*(1-Model!$N$48))"
            )
            if offset == MAX_YEARS - 1:
                next_pv = "0"
            else:
                next_pv = f"{get_column_letter(col + helper_width + 17)}{row}"
            proj[f"{pv_col}{row}"] = f'=IF(A{row}="","",{adjusted_need}+{next_pv}/(1+Model!$B$31))'

    for col in range(1, 19):
        proj.column_dimensions[get_column_letter(col)].width = 18
    for row in range(2, last_row + 1):
        for col in [3, 4, 5, 6, 7, 8, 10, 11, 12, 13, 14, 16, 17, 18]:
            money_style(proj.cell(row, col))
    proj.freeze_panes = "A2"
    proj.auto_filter.ref = f"A1:R{last_row}"

    chart = LineChart()
    chart.title = "Portefolje og slutsaldo"
    chart.y_axis.title = "Kr"
    chart.x_axis.title = "Alder"
    data = Reference(proj, min_col=3, max_col=8, min_row=1, max_row=last_row)
    cats = Reference(proj, min_col=1, min_row=2, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 12
    chart.width = 24
    proj.add_chart(chart, "L2")

    plan["A1"] = "Udbetalingsplan ved selvpensionsalder"
    plan["A1"].font = Font(size=16, bold=True)
    plan["A3"] = "Selvpensionsalder"
    plan["B3"] = "=Model!B28"
    plan["A4"] = "Startportefolje i plan"
    plan["B4"] = '=IF(ISNUMBER(B3),INDEX(\'Aarlig projektion\'!C:C,MATCH(B3,\'Aarlig projektion\'!A:A,0)),"")'
    money_style(plan["B4"])
    plan["A5"] = "Bemerkning"
    plan["B5"] = "Planen viser den simulerede brug af hvert depot ved selvpensionsalderen."

    plan_headers = [
        "Alder",
        "Holding start",
        "Pensionsdepot start",
        "Aktiedepot start",
        "ASK start",
        "Forbrug efter skat",
        "Indkomst efter skat",
        "Netto fra depoter",
        "Ejendom udbytte",
        "Ejendom loen",
        "Brutto pension",
        "Holding udbytte",
        "Holding loen",
        "Brutto aktiedepot",
        "Brutto ASK",
        "Mangel",
        "Slutsaldo samlet",
    ]
    for col, header in enumerate(plan_headers, start=1):
        plan.cell(7, col, header)
    style_header(plan[7])

    for row in range(8, last_row + 7):
        offset = row - 8
        if row == 8:
            plan[f"A{row}"] = '=IF(ISNUMBER($B$3),$B$3,"")'
        else:
            prev = row - 1
            plan[f"A{row}"] = f'=IF(OR(A{prev}="",A{prev}>=Model!$B$4),"",A{prev}+1)'

        source_row = 'MATCH($B$3,\'Aarlig projektion\'!$A:$A,0)'
        source_col = helper_start + offset * helper_width
        for plan_col, source_offset in zip(range(2, 18), [7, 8, 9, 10, None, None, None, 1, 2, 11, 12, 13, 14, 15, 21, None]):
            col_letter = get_column_letter(plan_col)
            if source_offset is None:
                continue
            source_letter = get_column_letter(source_col + source_offset)
            plan[f"{col_letter}{row}"] = f'=IF(A{row}="","",INDEX(\'Aarlig projektion\'!{source_letter}:{source_letter},{source_row}))'

        future_age = f"A{row}"
        target = f"IF({future_age}>Model!$B$4-Model!$B$8,Model!$B$7,Model!$B$6)"
        public_income = f"IF({future_age}>=Model!$B$10,Model!$B$23,0)"
        job_income = f"IF({offset}<Model!$B$15,Model!$B$25,0)"
        plan[f"F{row}"] = f'=IF(A{row}="","",{target})'
        plan[f"G{row}"] = f'=IF(A{row}="","",{public_income}+{job_income}+I{row}*(1-Model!$M$48)+J{row}*(1-Model!$N$48))'
        plan[f"H{row}"] = f'=IF(A{row}="","",MAX(0,F{row}-G{row}))'
        end_total_parts = [
            f"INDEX('Aarlig projektion'!{get_column_letter(source_col + i)}:{get_column_letter(source_col + i)},{source_row})"
            for i in range(16, 21)
        ]
        plan[f"Q{row}"] = f'=IF(A{row}="","",SUM({",".join(end_total_parts)}))'

    for col in range(1, 18):
        plan.column_dimensions[get_column_letter(col)].width = 18
    for row in range(8, last_row + 7):
        for col in range(2, 18):
            money_style(plan.cell(row, col))
    plan.freeze_panes = "A8"
    plan.auto_filter.ref = f"A7:Q{last_row + 6}"

    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
